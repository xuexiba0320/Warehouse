import json
import math

import requests
from fake_useragent import UserAgent  # 随机user—agent
from openpyxl import Workbook

# 可以使用， 会被检测爬虫进行ip拒绝访问
# url="https://api.bilibili.com/x/v2/reply?jsonp=jsonp&type=1&oid=458918620&pn="会缺部分数据
url = 'https://api.bilibili.com/x/v2/reply?jsonp=jsonp&type=1&oid=379648689&sort=2&pn='
USER_AGENT_object = UserAgent()
USER_AGENT = USER_AGENT_object.random
headers = {
    'user-agent': USER_AGENT
}


# 获取json数据
def download(url):
    print(headers)
    result = requests.get(url=url, headers=headers)
    print(result)
    return json.loads(result.text)["data"]


# excel数据导出
def excel_out(data):
    global index
    ws.cell(row=index, column=1, value=index - 1)
    ws.cell(row=index, column=2, value=data["member"]["uname"])
    ws.cell(row=index, column=3, value=data["member"]["level_info"]["current_level"])
    ws.cell(row=index, column=4, value=data["member"]["sex"])
    ws.cell(row=index, column=5, value=data["content"]["message"])
    ws.cell(row=index, column=6, value=data["like"])
    index += 1


index = 2;
wb = Workbook()
ws = wb.active
ws.title = "半吨仙人"
ws.cell(row=1, column=1, value="序号")
ws.cell(row=1, column=2, value="用户名")
ws.cell(row=1, column=3, value="等级")
ws.cell(row=1, column=4, value="性别")
ws.cell(row=1, column=5, value="评论")
ws.cell(row=1, column=6, value="点赞数")

data = download(url + "1")
totalPage = math.ceil(data["page"]["count"] / data["page"]["size"])
print("total: %s页" % totalPage)
for i in range(1, totalPage + 1):
    data = download(url + str(i))
    replies = data["replies"]
    for v in replies:
        excel_out(v)
    print("当前进度 %.2f%%" % (i / totalPage * 100))
wb.save("test.xlsx")
print("完成\（*.*）/")
